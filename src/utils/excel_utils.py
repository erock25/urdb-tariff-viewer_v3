"""
Excel export utilities for URDB Tariff Viewer.

This module contains functions for exporting tariff data to Excel format
with proper formatting for rates, percentages, and currency values.
"""

import io
import pandas as pd
import numpy as np
from typing import List, Optional
from datetime import datetime


def export_rate_table_to_excel(
    df: pd.DataFrame,
    sheet_name: str,
    rate_columns: List[str],
    percentage_columns: Optional[List[str]] = None,
    rate_precision: int = 4
) -> bytes:
    """
    Export a rate table DataFrame to Excel with proper formatting.
    
    Args:
        df: DataFrame to export (with formatted string values)
        sheet_name: Name of the Excel sheet
        rate_columns: Columns containing currency rates (e.g., "$0.1234")
        percentage_columns: Columns containing percentages (e.g., "50.0%")
        rate_precision: Number of decimal places for rate formatting
        
    Returns:
        Excel file content as bytes
    """
    from openpyxl.styles import numbers
    
    # Create a copy and convert formatted strings to numeric values
    excel_df = df.copy()
    
    # Convert rate columns (remove $ and convert to float)
    for col in rate_columns:
        if col in excel_df.columns:
            excel_df[col] = excel_df[col].astype(str).str.replace('$', '', regex=False).astype(float)
    
    # Convert percentage columns (remove % and convert to decimal)
    if percentage_columns:
        for col in percentage_columns:
            if col in excel_df.columns:
                excel_df[col] = excel_df[col].astype(str).str.replace('%', '', regex=False).astype(float) / 100
    
    # Create Excel file in memory
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        excel_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get worksheet for formatting
        worksheet = writer.sheets[sheet_name]
        headers = list(excel_df.columns)
        
        # Determine number format based on precision
        if rate_precision == 4:
            rate_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
        else:
            rate_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        # Apply formatting
        for col_idx, col_name in enumerate(headers, start=1):
            if col_name in rate_columns:
                for row_idx in range(2, len(excel_df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = rate_format
            elif percentage_columns and col_name in percentage_columns:
                for row_idx in range(2, len(excel_df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = '0.0%'
    
    return buffer.getvalue()


def apply_excel_rate_formatting(
    worksheet,
    df: pd.DataFrame,
    rate_columns: List[str],
    rate_precision: int = 4,
    start_row: int = 2
) -> None:
    """
    Apply currency formatting to rate columns in an Excel worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame being written
        rate_columns: Column names containing rate values
        rate_precision: Number of decimal places
        start_row: First data row (1-indexed, after header)
    """
    headers = list(df.columns)
    
    if rate_precision == 4:
        rate_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
    else:
        rate_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in rate_columns:
            for row_idx in range(start_row, len(df) + start_row):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = rate_format


def generate_energy_rates_excel(tariff_viewer, year: int = 2025) -> bytes:
    """
    Generate an Excel file with multiple sheets containing energy and demand rate data.
    
    Args:
        tariff_viewer: TariffViewer instance with energy and demand rate data
        year: Year for timeseries generation (default 2025)
        
    Returns:
        Excel file content as bytes
    """
    from openpyxl.styles import numbers
    from .rate_utils import generate_energy_rate_timeseries
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ENERGY RATES SHEETS
        _write_energy_rate_table(writer, tariff_viewer)
        _write_energy_rate_heatmaps(writer, tariff_viewer)
        _write_energy_timeseries(writer, tariff_viewer, year)
        
        # DEMAND RATES SHEETS
        _write_demand_rate_table(writer, tariff_viewer)
        _write_demand_rate_heatmaps(writer, tariff_viewer)
        _write_flat_demand_rates(writer, tariff_viewer)
    
    return output.getvalue()


def _write_energy_rate_table(writer, tariff_viewer) -> None:
    """Write energy rate table sheet."""
    try:
        tou_table = tariff_viewer.create_tou_labels_table()
        if not tou_table.empty:
            excel_tou = tou_table.copy()
            for col in ['Base Rate ($/kWh)', 'Adjustment ($/kWh)', 'Total Rate ($/kWh)']:
                if col in excel_tou.columns:
                    excel_tou[col] = excel_tou[col].str.replace('$', '').astype(float)
            if '% of Year' in excel_tou.columns:
                excel_tou['% of Year'] = excel_tou['% of Year'].str.replace('%', '').astype(float) / 100
            
            excel_tou.to_excel(writer, sheet_name='Energy Rate Table', index=False)
            
            worksheet = writer.sheets['Energy Rate Table']
            headers = list(excel_tou.columns)
            for col_idx, col_name in enumerate(headers, start=1):
                if col_name in ['Base Rate ($/kWh)', 'Adjustment ($/kWh)', 'Total Rate ($/kWh)']:
                    for row_idx in range(2, len(excel_tou) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
                elif col_name == '% of Year':
                    for row_idx in range(2, len(excel_tou) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '0.0%'
        else:
            no_data_df = pd.DataFrame({'Note': ['No energy rate structure found in tariff']})
            no_data_df.to_excel(writer, sheet_name='Energy Rate Table', index=False)
    except Exception as e:
        error_df = pd.DataFrame({'Error': [f'Could not generate rate table: {str(e)}']})
        error_df.to_excel(writer, sheet_name='Energy Rate Table', index=False)


def _write_energy_rate_heatmaps(writer, tariff_viewer) -> None:
    """Write weekday and weekend energy rate heatmap sheets."""
    # Weekday
    weekday_df = tariff_viewer.weekday_df.copy()
    weekday_df.columns = [f'{h:02d}:00' for h in range(24)]
    weekday_df.to_excel(writer, sheet_name='Weekday Energy Rates')
    
    worksheet = writer.sheets['Weekday Energy Rates']
    for row_idx in range(2, len(weekday_df) + 2):
        for col_idx in range(2, len(weekday_df.columns) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
    
    # Weekend
    weekend_df = tariff_viewer.weekend_df.copy()
    weekend_df.columns = [f'{h:02d}:00' for h in range(24)]
    weekend_df.to_excel(writer, sheet_name='Weekend Energy Rates')
    
    worksheet = writer.sheets['Weekend Energy Rates']
    for row_idx in range(2, len(weekend_df) + 2):
        for col_idx in range(2, len(weekend_df.columns) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'


def _write_energy_timeseries(writer, tariff_viewer, year: int) -> None:
    """Write full year energy timeseries sheet."""
    from .rate_utils import generate_energy_rate_timeseries
    
    timeseries_df = generate_energy_rate_timeseries(tariff_viewer, year)
    timeseries_df.to_excel(writer, sheet_name='Energy Timeseries', index=False)
    
    if 'energy_rate_$/kWh' in timeseries_df.columns:
        worksheet = writer.sheets['Energy Timeseries']
        rate_col_idx = list(timeseries_df.columns).index('energy_rate_$/kWh') + 1
        for row_idx in range(2, len(timeseries_df) + 2):
            cell = worksheet.cell(row=row_idx, column=rate_col_idx)
            cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'


def _write_demand_rate_table(writer, tariff_viewer) -> None:
    """Write demand rate table sheet."""
    try:
        demand_table = tariff_viewer.create_demand_labels_table()
        if not demand_table.empty:
            excel_demand = demand_table.copy()
            for col in ['Base Rate ($/kW)', 'Adjustment ($/kW)', 'Total Rate ($/kW)']:
                if col in excel_demand.columns:
                    excel_demand[col] = excel_demand[col].str.replace('$', '').astype(float)
            if '% of Year' in excel_demand.columns:
                excel_demand['% of Year'] = excel_demand['% of Year'].str.replace('%', '').astype(float) / 100
            
            excel_demand.to_excel(writer, sheet_name='Demand Rate Table', index=False)
            
            worksheet = writer.sheets['Demand Rate Table']
            headers = list(excel_demand.columns)
            for col_idx, col_name in enumerate(headers, start=1):
                if col_name in ['Base Rate ($/kW)', 'Adjustment ($/kW)', 'Total Rate ($/kW)']:
                    for row_idx in range(2, len(excel_demand) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                elif col_name == '% of Year':
                    for row_idx in range(2, len(excel_demand) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '0.0%'
        else:
            no_data_df = pd.DataFrame({'Note': ['No demand rate structure found in tariff']})
            no_data_df.to_excel(writer, sheet_name='Demand Rate Table', index=False)
    except Exception as e:
        error_df = pd.DataFrame({'Error': [f'Could not generate demand rate table: {str(e)}']})
        error_df.to_excel(writer, sheet_name='Demand Rate Table', index=False)


def _write_demand_rate_heatmaps(writer, tariff_viewer) -> None:
    """Write weekday and weekend demand rate heatmap sheets."""
    # Weekday
    demand_weekday_df = tariff_viewer.demand_weekday_df.copy()
    demand_weekday_df.columns = [f'{h:02d}:00' for h in range(24)]
    demand_weekday_df.to_excel(writer, sheet_name='Weekday Demand Rates')
    
    worksheet = writer.sheets['Weekday Demand Rates']
    for row_idx in range(2, len(demand_weekday_df) + 2):
        for col_idx in range(2, len(demand_weekday_df.columns) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Weekend
    demand_weekend_df = tariff_viewer.demand_weekend_df.copy()
    demand_weekend_df.columns = [f'{h:02d}:00' for h in range(24)]
    demand_weekend_df.to_excel(writer, sheet_name='Weekend Demand Rates')
    
    worksheet = writer.sheets['Weekend Demand Rates']
    for row_idx in range(2, len(demand_weekend_df) + 2):
        for col_idx in range(2, len(demand_weekend_df.columns) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def _write_flat_demand_rates(writer, tariff_viewer) -> None:
    """Write flat demand rates sheet."""
    flat_demand_df = tariff_viewer.flat_demand_df.copy()
    flat_demand_df.to_excel(writer, sheet_name='Flat Demand Rates')
    
    worksheet = writer.sheets['Flat Demand Rates']
    for row_idx in range(2, len(flat_demand_df) + 2):
        for col_idx in range(2, len(flat_demand_df.columns) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

